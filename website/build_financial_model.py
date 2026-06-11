"""Build ElderCare VN dynamic financial model (.xlsx) — formula-driven, ties to Phiên 1."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "public/eldercare-financial-model.xlsx"
FONT = "Arial"
BLUE = "0000FF"   # hardcoded inputs
BLACK = "000000"  # formulas
GREEN = "008000"  # cross-sheet links

YEARS = ["Năm 1", "Năm 2", "Năm 3", "Năm 4", "Năm 5"]
thin = Side(style="thin", color="D0D0D0")
border = Border(bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1A1A2E")
tot_fill = PatternFill("solid", fgColor="F0FDFA")
yellow = PatternFill("solid", fgColor="FFF7CC")
sub_fill = PatternFill("solid", fgColor="F4F6F8")

wb = Workbook()


def title(ws, text, ncol=6):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, bold=True, size=13, color="0D9488")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)


def note(ws, row, text, ncol=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, italic=True, size=8, color="6B7280")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)


def yearhdr(ws, row, label="", cols=YEARS):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name=FONT, bold=True, color="FFFFFF")
    c.fill = hdr_fill
    for i, y in enumerate(cols):
        cc = ws.cell(row=row, column=2 + i, value=y)
        cc.font = Font(name=FONT, bold=True, color="FFFFFF")
        cc.fill = hdr_fill
        cc.alignment = Alignment(horizontal="right")


def sethead(ws, row, text, ncol=6, fill="0D9488"):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, bold=True, color="FFFFFF")
    f = PatternFill("solid", fgColor=fill)
    for j in range(1, ncol + 1):
        ws.cell(row=row, column=j).fill = f


def put(ws, row, col, val, color=BLACK, fmt=None, bold=False, fill=None, align="right"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT, color=color, bold=bold)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if col > 1:
        c.alignment = Alignment(horizontal=align)
    return c


def label(ws, row, text, bold=False, indent=False):
    c = ws.cell(row=row, column=1, value=("   " + text if indent else text))
    c.font = Font(name=FONT, bold=bold)
    return c


NUM1 = '#,##0.0;(#,##0.0);"-"'
NUM0 = '#,##0;(#,##0);"-"'
PCT = '0.0%'
MULT = '0.0x'
MONEY = '#,##0.0'

# ============================ ASSUMPTIONS ============================
ws = wb.active
ws.title = "Assumptions"
title(ws, "ElderCare VN — Mô hình tài chính động · ASSUMPTIONS (giả định)")
note(ws, 2, "Ô chữ XANH = giả định nhập tay — đổi để chạy kịch bản; mọi sheet khác tính lại tự động. Đơn vị ghi trong nhãn. Số khớp 100% phần Tài chính (Phiên 1).")

r = 4
yearhdr(ws, r, "Giả định vận hành")
op = {}
data_op = [
    ("Gia đình trả phí (nghìn)", [6, 30, 110, 280, 520], NUM0),
    ("Net ARPU (nghìn VND/tháng)", [110, 125, 140, 150, 152], NUM0),
    ("Gross margin (%)", [0.52, 0.56, 0.60, 0.62, 0.64], PCT),
]
r += 1
for name, vals, fmt in data_op:
    label(ws, r, name)
    for i, v in enumerate(vals):
        put(ws, r, 2 + i, v, BLUE, fmt)
    op[name] = r
    r += 1

r += 1
sethead(ws, r, "Chi phí hoạt động (tỷ VND)")
r += 1
cost = {}
data_cost = [
    ("Nhân sự (lương loaded)", [12.8, 37.4, 98.0, 201.6, 351.5]),
    ("Marketing (S&M)", [11.0, 30.0, 64.0, 110.0, 130.0]),
    ("Công nghệ (cloud/AI/license)", [4.0, 8.0, 18.0, 32.0, 45.0]),
    ("Vận hành & G&A", [5.0, 12.0, 28.0, 45.0, 50.0]),
    ("Khấu hao (D&A)", [1.0, 3.0, 7.0, 14.0, 22.0]),
]
for name, vals in data_cost:
    label(ws, r, name)
    for i, v in enumerate(vals):
        put(ws, r, 2 + i, v, BLUE, NUM1)
    cost[name] = r
    r += 1

r += 1
sethead(ws, r, "Unit economics (steady-state)")
r += 1
ue = {}
data_ue = [
    ("Net ARPU/tháng (nghìn VND)", 150, NUM0),
    ("GM blended (%)", 0.62, PCT),
    ("CAC blended (nghìn VND)", 480, NUM0),
    ("Churn/tháng (%)", 0.033, PCT),
]
for name, v, fmt in data_ue:
    label(ws, r, name)
    put(ws, r, 2, v, BLUE, fmt)
    ue[name] = r
    r += 1

r += 1
sethead(ws, r, "Tài chính & định giá")
r += 1
fin = {}


def fin_row(name, v, fmt, color=BLUE, formula=False):
    global r
    label(ws, r, name)
    put(ws, r, 2, v, color, fmt)
    fin[name] = r
    r += 1


fin_row("Tỷ giá (tỷ VND / triệu USD)", 25.5, NUM1)
fin_row("Thuế TNDN (%)", 0.20, PCT)
fin_row("Rf — TPCP VN 10 năm (%)", 0.045, PCT)
fin_row("ERP Việt Nam — Damodaran (%)", 0.0813, PCT)
fin_row("Beta (Healthcare Info & Tech)", 1.11, '0.00')
# Ke formula
label(ws, r, "Ke = Rf + Beta × ERP (%)")
put(ws, r, 2, f"=B{fin['Rf — TPCP VN 10 năm (%)']}+B{fin['Beta (Healthcare Info & Tech)']}*B{fin['ERP Việt Nam — Damodaran (%)']}", BLACK, PCT)
fin["Ke"] = r
r += 1
fin_row("Phần bù quy mô/giai đoạn (%)", 0.065, PCT)
label(ws, r, "WACC / suất chiết khấu (%)", bold=True)
put(ws, r, 2, f"=B{fin['Ke']}+B{fin['Phần bù quy mô/giai đoạn (%)']}", BLACK, PCT, bold=True, fill=tot_fill)
ws.cell(row=r, column=1).fill = tot_fill
fin["WACC"] = r
r += 1
fin_row("Exit multiple (× DT Y5)", 5.0, MULT)

ws.column_dimensions["A"].width = 34
for col in "BCDEF":
    ws.column_dimensions[col].width = 13

A = {"op": op, "cost": cost, "ue": ue, "fin": fin}

# ============================ REVENUE BUILD ============================
ws = wb.create_sheet("Revenue build")
title(ws, "Doanh thu — build từ cohort & 5 trụ cột")
note(ws, 2, "Doanh thu thuần = Gia đình × ARPU × 12. Tách 5 trụ cột theo % cơ cấu Y5 (ước tính nội bộ). Link XANH LÁ kéo từ sheet Assumptions.")
r = 4
yearhdr(ws, r, "Cohort")
r += 1
label(ws, r, "Gia đình trả phí (nghìn)")
for i in range(5):
    put(ws, r, 2 + i, f"=Assumptions!{get_column_letter(2+i)}{op['Gia đình trả phí (nghìn)']}", GREEN, NUM0)
fam_row = r
r += 1
label(ws, r, "Net ARPU (nghìn VND/tháng)")
for i in range(5):
    put(ws, r, 2 + i, f"=Assumptions!{get_column_letter(2+i)}{op['Net ARPU (nghìn VND/tháng)']}", GREEN, NUM0)
arpu_row = r
r += 1
label(ws, r, "Doanh thu thuần (tỷ VND)", bold=True)
for i in range(5):
    cl = get_column_letter(2 + i)
    put(ws, r, 2 + i, f"={cl}{fam_row}*{cl}{arpu_row}*12/1000", BLACK, NUM1, bold=True, fill=tot_fill)
ws.cell(row=r, column=1).fill = tot_fill
rev_row = r
r += 2

sethead(ws, r, "Tách 5 trụ cột (% cơ cấu DT net — Y5)")
r += 1
yearhdr(ws, r, "Trụ cột", cols=["% DT", "Take-rate / cơ chế", "Gross margin", "DT Y5 (tỷ)", ""])
r += 1
pillars = [
    ("CareConnect (điều dưỡng)", 0.36, "Hoa hồng 20% GMV", 0.85),
    ("Khám & Telehealth", 0.25, "Sub 299–599k + 40% phí khám", 0.60),
    ("MedFin (tài chính y tế)", 0.20, "NIM 6–9% + hoa hồng BH", 0.55),
    ("CareSense (thiết bị)", 0.12, "Thuê 99–199k/th + SaaS", 0.45),
    ("Dược & thiết bị", 0.07, "Markup 12–18%", 0.25),
]
pillar_start = r
for name, pct, mech, gm in pillars:
    label(ws, r, name)
    put(ws, r, 2, pct, BLUE, PCT)
    put(ws, r, 3, mech, BLACK, align="left")
    put(ws, r, 4, gm, BLUE, PCT)
    put(ws, r, 5, f"=B{r}*F{rev_row}", BLACK, NUM1)
    r += 1
label(ws, r, "Tổng / blended", bold=True)
put(ws, r, 2, f"=SUM(B{pillar_start}:B{r-1})", BLACK, PCT, bold=True, fill=tot_fill)
put(ws, r, 4, f"=SUMPRODUCT(B{pillar_start}:B{r-1},D{pillar_start}:D{r-1})", BLACK, PCT, bold=True, fill=tot_fill)
put(ws, r, 5, f"=SUM(E{pillar_start}:E{r-1})", BLACK, NUM1, bold=True, fill=tot_fill)
ws.cell(row=r, column=1).fill = tot_fill
note(ws, r + 2, "Kiểm tra: tổng % = 100% và tổng DT trụ cột = Doanh thu thuần Y5 (~948 tỷ). GM blended trụ cột ≈ 62–64% khớp Assumptions.")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["C"].width = 30
for col in "BDEF":
    ws.column_dimensions[col].width = 14
REV = {"rev_row": rev_row}

# ============================ P&L ============================
ws = wb.create_sheet("P&L 5 năm")
title(ws, "Dự phóng P&L 5 năm (tỷ VND)")
note(ws, 2, "Link XANH LÁ từ Assumptions / Revenue build. EBITDA, EBIT, thuế (NOL), lợi nhuận = công thức đen. Thuế ≈ 0 khi lỗ lũy kế còn âm.")
r = 4
yearhdr(ws, r, "Chỉ tiêu (tỷ VND)")
r += 1
pl = {}


def pl_line(name, builder, fmt=NUM1, bold=False, fill=None, color=BLACK, indent=False):
    global r
    pl[name] = r
    label(ws, r, name, bold=bold, indent=indent)
    if fill:
        ws.cell(row=r, column=1).fill = fill
    for i in range(5):
        cl = get_column_letter(2 + i)
        put(ws, r, 2 + i, builder(cl, i), color, fmt, bold=bold, fill=fill)
    r += 1


pl_line("Doanh thu thuần", lambda cl, i: f"='Revenue build'!{cl}{rev_row}", color=GREEN, bold=True)
pl_line("Gross margin (%)", lambda cl, i: f"=Assumptions!{cl}{op['Gross margin (%)']}", fmt=PCT, color=GREEN)
pl_line("Giá vốn (COGS)", lambda cl, i: f"={cl}{pl['Doanh thu thuần']}*(1-{cl}{pl['Gross margin (%)']})")
pl_line("Lợi nhuận gộp", lambda cl, i: f"={cl}{pl['Doanh thu thuần']}*{cl}{pl['Gross margin (%)']}", bold=True, fill=tot_fill)
pl_line("— Nhân sự", lambda cl, i: f"=Assumptions!{cl}{cost['Nhân sự (lương loaded)']}", color=GREEN, indent=True)
pl_line("— Marketing", lambda cl, i: f"=Assumptions!{cl}{cost['Marketing (S&M)']}", color=GREEN, indent=True)
pl_line("— Công nghệ", lambda cl, i: f"=Assumptions!{cl}{cost['Công nghệ (cloud/AI/license)']}", color=GREEN, indent=True)
pl_line("— Vận hành & G&A", lambda cl, i: f"=Assumptions!{cl}{cost['Vận hành & G&A']}", color=GREEN, indent=True)
pl_line("Tổng OpEx", lambda cl, i: f"=SUM({cl}{pl['— Nhân sự']}:{cl}{pl['— Vận hành & G&A']})", bold=True)
pl_line("EBITDA", lambda cl, i: f"={cl}{pl['Lợi nhuận gộp']}-{cl}{pl['Tổng OpEx']}", bold=True, fill=tot_fill)
pl_line("Khấu hao (D&A)", lambda cl, i: f"=Assumptions!{cl}{cost['Khấu hao (D&A)']}", color=GREEN, indent=True)
pl_line("EBIT", lambda cl, i: f"={cl}{pl['EBITDA']}-{cl}{pl['Khấu hao (D&A)']}", bold=True)
# cumulative EBIT for NOL
pl_line("EBIT lũy kế", lambda cl, i: (f"={cl}{pl['EBIT']}" if i == 0 else f"={get_column_letter(1+i)}{pl['EBIT lũy kế']}+{cl}{pl['EBIT']}"), color="6B7280", indent=True)
pl_line("Thuế TNDN (NOL)", lambda cl, i: f"=IF({cl}{pl['EBIT lũy kế']}>0,{cl}{pl['EBIT']}*Assumptions!$B${fin['Thuế TNDN (%)']},0)", indent=True)
pl_line("Lợi nhuận sau thuế", lambda cl, i: f"={cl}{pl['EBIT']}-{cl}{pl['Thuế TNDN (NOL)']}", bold=True, fill=tot_fill)
pl_line("Lỗ lũy kế", lambda cl, i: (f"={cl}{pl['Lợi nhuận sau thuế']}" if i == 0 else f"={get_column_letter(1+i)}{pl['Lỗ lũy kế']}+{cl}{pl['Lợi nhuận sau thuế']}"), color="6B7280", indent=True)
ws.column_dimensions["A"].width = 26
for col in "BCDEF":
    ws.column_dimensions[col].width = 12
PL = pl

# ============================ UNIT ECONOMICS ============================
ws = wb.create_sheet("Unit economics")
title(ws, "Unit economics (trên 1 gia đình trả phí)")
note(ws, 2, "Tất cả tính từ Assumptions (link XANH LÁ) → công thức đen. LTV theo lợi nhuận gộp (thận trọng).")
r = 4
uerows = {}


def ue_line(name, formula, fmt, color=BLACK, src=None, bold=False):
    global r
    label(ws, r, name, bold=bold)
    put(ws, r, 2, formula, color, fmt, bold=bold)
    if src:
        put(ws, r, 3, src, "6B7280", align="left")
    uerows[name] = r
    r += 1


ue_line("Net ARPU/tháng (nghìn VND)", f"=Assumptions!B{ue['Net ARPU/tháng (nghìn VND)']}", NUM0, GREEN)
ue_line("GM blended (%)", f"=Assumptions!B{ue['GM blended (%)']}", PCT, GREEN)
ue_line("Lợi nhuận gộp/tháng (nghìn)", f"=B{uerows['Net ARPU/tháng (nghìn VND)']}*B{uerows['GM blended (%)']}", NUM0, src="ARPU × GM")
ue_line("CAC blended (nghìn VND)", f"=Assumptions!B{ue['CAC blended (nghìn VND)']}", NUM0, GREEN)
ue_line("Churn/tháng (%)", f"=Assumptions!B{ue['Churn/tháng (%)']}", PCT, GREEN)
ue_line("Tuổi thọ KH (tháng)", f"=1/B{uerows['Churn/tháng (%)']}", NUM1, src="1 / churn")
ue_line("Payback CAC (tháng)", f"=B{uerows['CAC blended (nghìn VND)']}/B{uerows['Lợi nhuận gộp/tháng (nghìn)']}", NUM1, src="CAC / LN gộp tháng", bold=True)
ue_line("LTV (nghìn VND)", f"=B{uerows['Lợi nhuận gộp/tháng (nghìn)']}*B{uerows['Tuổi thọ KH (tháng)']}", NUM0, src="LN gộp tháng × tuổi thọ", bold=True)
ue_line("LTV / CAC", f"=B{uerows['LTV (nghìn VND)']}/B{uerows['CAC blended (nghìn VND)']}", MULT, src="mục tiêu ≥ 3×", bold=True)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 26

# ============================ FUNDING & CAP TABLE ============================
ws = wb.create_sheet("Funding & cap table")
title(ws, "Gọi vốn 4 vòng & cap table (pha loãng)")
note(ws, 2, "Số tiền vòng (xanh) → $M, post-money, % pha loãng tính tự động. Cap table dùng % pha loãng để cuốn chiếu qua các vòng. Link XANH LÁ = tỷ giá từ Assumptions.")
r = 4
c = ws.cell(row=r, column=1, value="Vòng")
c.font = Font(name=FONT, bold=True, color="FFFFFF"); c.fill = hdr_fill
for j, h in enumerate(["Số tiền (tỷ)", "Quy mô ($M)", "Pre-money ($M)", "Post-money ($M)", "% pha loãng"]):
    cc = ws.cell(row=r, column=2 + j, value=h)
    cc.font = Font(name=FONT, bold=True, color="FFFFFF"); cc.fill = hdr_fill
    cc.alignment = Alignment(horizontal="right")
r += 1
rounds = [("Seed", 38, 6), ("Series A", 150, 18), ("Series B", 510, 70), ("Series C", 1275, 250)]
round_start = r
dil_rows = {}
for name, amt, pre in rounds:
    label(ws, r, name, bold=True)
    put(ws, r, 2, amt, BLUE, NUM0)
    put(ws, r, 3, f"=B{r}/Assumptions!$B${fin['Tỷ giá (tỷ VND / triệu USD)']}", GREEN if False else BLACK, NUM1)
    put(ws, r, 4, pre, BLUE, NUM1)
    put(ws, r, 5, f"=D{r}+C{r}", BLACK, NUM1)
    put(ws, r, 6, f"=C{r}/E{r}", BLACK, PCT)
    dil_rows[name] = r
    r += 1
r += 2

# Cap table
sethead(ws, r, "Cap table — % sở hữu sau mỗi vòng", ncol=6)
r += 1
c = ws.cell(row=r, column=1, value="Cổ đông")
c.font = Font(name=FONT, bold=True, color="FFFFFF"); c.fill = hdr_fill
for j, h in enumerate(["Khởi đầu", "Sau Seed", "Sau Series A", "Sau Series B", "Sau Series C"]):
    cc = ws.cell(row=r, column=2 + j, value=h)
    cc.font = Font(name=FONT, bold=True, color="FFFFFF"); cc.fill = hdr_fill
    cc.alignment = Alignment(horizontal="right")
r += 1
# dilution multiplier per round col: col C(seed)=row dil seed etc.
dseed = f"$F${dil_rows['Seed']}"
da = f"$F${dil_rows['Series A']}"
db = f"$F${dil_rows['Series B']}"
dc = f"$F${dil_rows['Series C']}"
cap_start = r
# Founders & ESOP: start value, then *(1-dil)
def carry_row(name, start_val):
    global r
    label(ws, r, name)
    put(ws, r, 2, start_val, BLUE, PCT)
    put(ws, r, 3, f"=B{r}*(1-{dseed})", BLACK, PCT)
    put(ws, r, 4, f"=C{r}*(1-{da})", BLACK, PCT)
    put(ws, r, 5, f"=D{r}*(1-{db})", BLACK, PCT)
    put(ws, r, 6, f"=E{r}*(1-{dc})", BLACK, PCT)
    r += 1

carry_row("Nhà sáng lập (6 founders)", 0.88)
carry_row("ESOP (pool nhân sự)", 0.12)
# Seed investors: appears at Sau Seed = dil seed, then diluted
label(ws, r, "Seed investors")
put(ws, r, 2, "", BLACK, PCT)
put(ws, r, 3, f"={dseed}", BLACK, PCT)
put(ws, r, 4, f"=C{r}*(1-{da})", BLACK, PCT)
put(ws, r, 5, f"=D{r}*(1-{db})", BLACK, PCT)
put(ws, r, 6, f"=E{r}*(1-{dc})", BLACK, PCT)
r += 1
label(ws, r, "Series A investors")
for col in (2, 3):
    put(ws, r, col, "", BLACK, PCT)
put(ws, r, 4, f"={da}", BLACK, PCT)
put(ws, r, 5, f"=D{r}*(1-{db})", BLACK, PCT)
put(ws, r, 6, f"=E{r}*(1-{dc})", BLACK, PCT)
r += 1
label(ws, r, "Series B investors")
for col in (2, 3, 4):
    put(ws, r, col, "", BLACK, PCT)
put(ws, r, 5, f"={db}", BLACK, PCT)
put(ws, r, 6, f"=E{r}*(1-{dc})", BLACK, PCT)
r += 1
label(ws, r, "Series C investors")
for col in (2, 3, 4, 5):
    put(ws, r, col, "", BLACK, PCT)
put(ws, r, 6, f"={dc}", BLACK, PCT)
r += 1
label(ws, r, "Tổng", bold=True)
for col in range(2, 7):
    cl = get_column_letter(col)
    put(ws, r, col, f"=SUM({cl}{cap_start}:{cl}{r-1})", BLACK, PCT, bold=True, fill=tot_fill)
ws.cell(row=r, column=1).fill = tot_fill
note(ws, r + 2, "Founders ~34% sau Series C. Mô hình pha loãng pro-rata đơn giản hóa (chưa tính ESOP refresh & anti-dilution). Tổng mỗi cột = 100%.")
ws.column_dimensions["A"].width = 28
for col in "BCDEF":
    ws.column_dimensions[col].width = 14

# ============================ DCF & SENSITIVITY ============================
ws = wb.create_sheet("DCF & sensitivity")
title(ws, "Định giá — DCF (CAPM/WACC) + sensitivity")
note(ws, 2, "WACC, exit multiple, DT Y5 link từ Assumptions/P&L. FCF rõ ràng (xanh) = ước tính. EV = ΣPV(FCF) + PV(terminal). Sensitivity tính trực tiếp bằng công thức (không hardcode).")
r = 4
yearhdr(ws, r, "DCF base (tỷ VND)")
r += 1
label(ws, r, "Năm (t)")
for i in range(5):
    put(ws, r, 2 + i, i + 1, BLACK, NUM0)
tnum_row = r
r += 1
label(ws, r, "FCF (≈EBITDA−capex−ΔNWC)")
base_fcf = [-31, -66, -102, -82, 18]
for i, v in enumerate(base_fcf):
    put(ws, r, 2 + i, v, BLUE, NUM0)
fcf_row = r
r += 1
label(ws, r, "Hệ số chiết khấu")
for i in range(5):
    cl = get_column_letter(2 + i)
    put(ws, r, 2 + i, f"=1/(1+Assumptions!$B${fin['WACC']})^{cl}{tnum_row}", BLACK, '0.000')
df_row = r
r += 1
label(ws, r, "PV của FCF")
for i in range(5):
    cl = get_column_letter(2 + i)
    put(ws, r, 2 + i, f"={cl}{fcf_row}*{cl}{df_row}", BLACK, NUM1)
pvfcf_row = r
r += 2

dcf = {}
def kv(name, formula, fmt, color=BLACK, bold=False):
    global r
    label(ws, r, name, bold=bold)
    put(ws, r, 2, formula, color, fmt, bold=bold, fill=tot_fill if bold else None)
    if bold:
        ws.cell(row=r, column=1).fill = tot_fill
    dcf[name] = r
    r += 1

kv("PV tổng FCF Y1–5 (tỷ)", f"=SUM(B{pvfcf_row}:F{pvfcf_row})", NUM1)
kv("DT Y5 (tỷ)", f"='P&L 5 năm'!F{PL['Doanh thu thuần']}", NUM1, GREEN)
kv("Exit multiple", f"=Assumptions!B{fin['Exit multiple (× DT Y5)']}", MULT, GREEN)
kv("Terminal value Y5 (tỷ)", f"=B{dcf['DT Y5 (tỷ)']}*B{dcf['Exit multiple']}", NUM1)
kv("PV terminal value (tỷ)", f"=B{dcf['Terminal value Y5 (tỷ)']}*F{df_row}", NUM1)
kv("EV hôm nay (tỷ VND)", f"=B{dcf['PV tổng FCF Y1–5 (tỷ)']}+B{dcf['PV terminal value (tỷ)']}", NUM1, bold=True)
kv("EV ($M)", f"=B{dcf['EV hôm nay (tỷ VND)']}/Assumptions!B{fin['Tỷ giá (tỷ VND / triệu USD)']}", NUM1, bold=True)
r += 1

# Sensitivity 2-var: exit (rows) x WACC (cols)
sethead(ws, r, "Sensitivity — EV (tỷ VND): Exit multiple × WACC", ncol=6)
r += 1
waccs = [0.16, 0.18, 0.20, 0.22, 0.24]
label(ws, r, "Exit ↓ / WACC →")
for j, w in enumerate(waccs):
    put(ws, r, 2 + j, w, BLUE, PCT)
sens_wacc_row = r
r += 1
exits = [3, 4, 5, 6, 7]
sens_start = r
fcf_rng = f"$B${fcf_row}:$F${fcf_row}"
t_rng = f"$B${tnum_row}:$F${tnum_row}"
for ex in exits:
    put(ws, r, 1, ex, BLUE, MULT)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="right")
    for j in range(5):
        wcell = f"{get_column_letter(2+j)}${sens_wacc_row}"
        cl = get_column_letter(2 + j)
        # EV = SUMPRODUCT(FCF, 1/(1+w)^t) + DTY5*exit/(1+w)^5
        f = (f"=SUMPRODUCT({fcf_rng},1/(1+{wcell})^{t_rng})"
             f"+$B${dcf['DT Y5 (tỷ)']}*$A{r}/(1+{wcell})^5")
        put(ws, r, 2 + j, f, BLACK, NUM0)
    r += 1
note(ws, r, "Ô tô đậm cột giữa (WACC 20%, exit 5×) = kịch bản base ≈ 1.740 tỷ. Lưới tính trực tiếp — không có giá trị hardcode.")
r += 2

# Base / Bull / Bear scenarios (each with own FCF & assumptions -> dynamic EV)
sethead(ws, r, "Kịch bản Bear / Base / Bull (EV động)", ncol=4)
r += 1
hdr = ws.cell(row=r, column=1, value="Giả định"); hdr.font = Font(name=FONT, bold=True, color="FFFFFF"); hdr.fill = hdr_fill
for j, h in enumerate(["🐻 Bear", "🎯 Base", "🚀 Bull"]):
    cc = ws.cell(row=r, column=2 + j, value=h); cc.font = Font(name=FONT, bold=True, color="FFFFFF"); cc.fill = hdr_fill
    cc.alignment = Alignment(horizontal="right")
r += 1
sc = {}
# rows: DT Y5, churn(info), exit, WACC, FCF1..5, EV tỷ, EV $M
def sc_row(name, bear, base, bull, fmt, color=BLUE):
    global r
    label(ws, r, name)
    put(ws, r, 2, bear, color, fmt)
    put(ws, r, 3, base, color, fmt)
    put(ws, r, 4, bull, color, fmt)
    sc[name] = r
    r += 1

sc_row("DT Y5 (tỷ)", 600, f"='P&L 5 năm'!F{PL['Doanh thu thuần']}", 1400, NUM0)
# fix base color to green for the linked cell
ws.cell(row=sc["DT Y5 (tỷ)"], column=3).font = Font(name=FONT, color=GREEN)
sc_row("Churn/tháng (%)", 0.05, 0.033, 0.02, PCT)
sc_row("Exit multiple", 3.0, 5.0, 7.0, MULT)
sc_row("WACC (%)", 0.24, 0.20, 0.18, PCT)
fcf_label = ["FCF Năm 1", "FCF Năm 2", "FCF Năm 3", "FCF Năm 4", "FCF Năm 5"]
bear_fcf = [-32, -66, -100, -80, 10]
base_fcf_s = [-31, -66, -102, -82, 18]
bull_fcf = [-28, -55, -80, -40, 40]
sc_fcf_start = r
for i in range(5):
    sc_row(fcf_label[i], bear_fcf[i], base_fcf_s[i], bull_fcf[i], NUM0)
    # vertical year helper in column F (for SUMPRODUCT dimension match)
    put(ws, r - 1, 6, i + 1, "6B7280", NUM0)
sc_fcf_end = r - 1
sc_t_rng = f"$F${sc_fcf_start}:$F${sc_fcf_end}"
# EV row using SUMPRODUCT per scenario (FCF vertical × year vertical)
label(ws, r, "EV (tỷ VND)", bold=True)
for j, col in enumerate("BCD"):
    fcf_c = f"{col}{sc_fcf_start}:{col}{sc_fcf_end}"
    w = f"{col}{sc['WACC (%)']}"
    f = (f"=SUMPRODUCT({fcf_c},1/(1+{w})^{sc_t_rng})"
         f"+{col}{sc['DT Y5 (tỷ)']}*{col}{sc['Exit multiple']}/(1+{w})^5")
    put(ws, r, 2 + j, f, BLACK, NUM0, bold=True, fill=tot_fill)
ws.cell(row=r, column=1).fill = tot_fill
sc["EV"] = r
r += 1
label(ws, r, "EV ($M)", bold=True)
for j, col in enumerate("BCD"):
    put(ws, r, 2 + j, f"={col}{sc['EV']}/Assumptions!$B${fin['Tỷ giá (tỷ VND / triệu USD)']}", BLACK, NUM1, bold=True, fill=tot_fill)
ws.cell(row=r, column=1).fill = tot_fill
note(ws, r + 2, "EV mỗi kịch bản tính từ chính FCF & giả định của kịch bản đó (động). Base ≈ 1.740 tỷ ($68M); Bear ≈ 460 tỷ ($18M); Bull ≈ 4.160 tỷ ($163M).")
ws.column_dimensions["A"].width = 26
for col in "BCDEF":
    ws.column_dimensions[col].width = 13

try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass
wb.save(OUT)
print("saved", OUT)
